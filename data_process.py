import pandas as pd
import openpyxl
import json
import xlsxwriter
import re
import numpy as np

# Helper function: Convert Excel column names to DataFrame column indices
def excel_col_to_num(col):
    num = 0
    for c in col:
        if 'A' <= c <= 'Z':
            num = num * 26 + (ord(c) - ord('A')) + 1
    return num - 1  # Column indices start from 0

# Part 1: Read formulas from the 5th row starting from the HJ column in an Excel file and save them to JSON
def save_formulas_to_json(excel_path, json_path, start_row=5, start_col='HJ'):
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb['Data']
    formulas = {}

    # Get the starting column index
    start_col_index = openpyxl.utils.column_index_from_string(start_col)

    for row in ws.iter_rows(min_row=start_row, max_row=start_row):
        for cell in row[start_col_index - 1:]:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[cell.coordinate] = cell.value

    # Modify the specific formula for "IE5"
    formulas["IE5"] = "=(C5-32)/1.8"

    with open(json_path, 'w') as f:
        json.dump(formulas, f, indent=4)

# Helper function to apply formula for each row
def apply_formula(row, formula, col_to_idx, df):
    def replace_cell(match):
        cell = match.group(0)
        if cell.startswith('$'):
            abs_col = re.search(r'\$([A-Z]+)', cell).group(1)
            abs_row = re.search(r'\$([0-9]+)', cell).group(1)
            col_idx = col_to_idx.get(abs_col, None)
            if col_idx is not None:
                return f'df.iloc[{int(abs_row) - 1}, {col_idx}]'
            else:
                return cell
        col = re.match(r'([A-Z]+)', cell).group(0)
        row_idx = row.name
        col_idx = col_to_idx.get(col, None)
        if col_idx is not None:
            return f'df.iloc[{row_idx}, {col_idx}]'
        else:
            return cell

    # Handle EXP() function in the formula
    def replace_exp(match):
        exp_content = match.group(1)
        return f'np.exp({exp_content})'

    # Handle LN() function in the formula
    def replace_ln(match):
        ln_content = match.group(1)
        return f'np.log({ln_content})'

    # Handle SQRT() function in the formula
    def replace_sqrt(match):
        sqrt_content = match.group(1)
        return f'np.sqrt({sqrt_content})'

    # Replace cell references in the formula
    formula_str = re.sub(r'\$?[A-Z]+\$?[0-9]+', replace_cell, formula[1:])

    # Replace EXP() function in the formula
    formula_str = re.sub(r'EXP\(([^)]+)\)', replace_exp, formula_str)

    # Replace LN() function in the formula
    formula_str = re.sub(r'LN\(([^)]+)\)', replace_ln, formula_str)

    # Replace SQRT() function in the formula
    formula_str = re.sub(r'SQRT\(([^)]+)\)', replace_sqrt, formula_str)

    # Replace ^ with **
    formula_str = formula_str.replace('^', '**')

    try:
        result = eval(formula_str, {"__builtins__": None}, {"df": df, "np": np})
    except ZeroDivisionError:
        result = 99  # Return 99 if division by zero
    except Exception as e:
        result = 99  # Return 99 if formula evaluation fails
    return result

# Part 2: Read data and process it
def process_data(sample_path, raw_data_path, output_path, formulas_json, df_final_path, address_path):
    # Read data from columns A to HJ in raw_data.xlsx starting from the third row
    df_raw_data = pd.read_excel(raw_data_path, sheet_name='Sheet1', usecols="A:HJ", skiprows=2, header=None)

    # Read data from sample.xlsx
    wb = openpyxl.load_workbook(sample_path, data_only=False)
    ws = wb['Data']
    df_sample = pd.DataFrame(ws.values)

    # Save all non-formula values after column HK and from the fifth row (inclusive)
    for row in range(0, df_sample.shape[0]):
        for col in range(excel_col_to_num('HK'), df_sample.shape[1]):
            cell_value = df_sample.iloc[row, col]
            if not (isinstance(cell_value, str) and cell_value.startswith('=')):
                df_sample.iloc[row, col] = cell_value

    # Copy data from A3 and A4 in raw_data.xlsx to the corresponding positions in df_sample
    df_sample.iloc[2, 1] = df_raw_data.iloc[0, 1]  # B3
    df_sample.iloc[3, 1] = df_raw_data.iloc[1, 1]  # B4

    # Read saved formulas
    with open(formulas_json, 'r') as f:
        formulas = json.load(f)

    # Create a mapping from Excel column names to DataFrame column indices
    col_to_idx = {}
    for i in range(702):
        col = ''
        temp = i
        while temp >= 0:
            col = chr(temp % 26 + ord('A')) + col
            temp = temp // 26 - 1
        col_to_idx[col] = i

    # Dynamically adjust the size of df_extended to accommodate all data
    max_rows = max(len(df_sample), len(df_raw_data)) + 2  # +2 for extra space
    num_cols = max(600, df_sample.shape[1], df_raw_data.shape[1])  # Ensure enough columns
    df_extended = pd.DataFrame(index=range(max_rows), columns=range(num_cols))
    df_extended.iloc[2:df_raw_data.shape[0] + 2, :df_raw_data.shape[1]] = df_raw_data.values

    # Record i-5 in the HL column of df_extended starting from the fifth row
    for i in range(4, len(df_extended) + 5):
        second_col_value = df_extended.iloc[i, 1]  # Get the value of the second column in the current row
        if pd.isna(second_col_value) or second_col_value == 0:
            break
        df_extended.iloc[i, col_to_idx['HL']] = i - 4  # Record the value in the HL column

    # If there is a number instead of a formula in the sample.xlsx file after column HK, copy the number to df_extended
    for row in range(0, df_sample.shape[0]):
        for col in range(col_to_idx['HK'], df_sample.shape[1]):
            cell_value = df_sample.iloc[row, col]
            if pd.notna(cell_value) and not isinstance(cell_value, str) and not str(cell_value).startswith('='):
                df_extended.iloc[row, col] = cell_value

    # Ensure previously copied values are not overwritten by 0
    df_extended = df_extended.mask(df_extended.isna(), 0)

    # Find the row index of the last non-zero or non-NaN cell from row=4
    for idx in reversed(range(4, len(df_extended))):
        if pd.notna(df_extended.iloc[idx, 1]) and df_extended.iloc[idx, 1] != 0:
            last_row_df_extended = idx
            break
    else:
        last_row_df_extended = 3  # If not found, set to 3 to start from row=4

    # Clear and set to 0 all data after last_row_df_extended
    df_extended.iloc[last_row_df_extended + 1:] = 0

    # Calculate the average value of each column in df_raw_data (starting from row=2 to last_row_df_extended), and place the average value in row last_row_df_extended + 1
    for col in range(2, df_raw_data.shape[1]):
        numeric_col = pd.to_numeric(df_extended.iloc[4:last_row_df_extended + 1, col], errors='coerce').fillna(0)
        avg_value = numeric_col.mean()
        df_extended.loc[last_row_df_extended + 1, col] = avg_value

    # Apply formulas to df_extended starting from the fifth row until any required parameter cell in the formula is missing data
    for coord, formula in formulas.items():
        col = re.match(r'([A-Z]+)', coord).group(0)  # Extract the column name part
        if col not in col_to_idx:
            continue

        start_row = int(re.match(r'[A-Z]+([0-9]+)', coord).group(1)) - 5  # Convert row number to 0-indexed
        if start_row < 0 or start_row >= len(df_extended):
            continue

        col_idx = col_to_idx[col]
        if col_idx >= df_extended.shape[1]:
            continue  # Ensure no out-of-bounds

        for row_num in range(start_row, len(df_extended)):
            if row_num < 4:  # Apply formula only to rows >= 4
                continue
            row = df_extended.iloc[row_num]

            # Check if the value in column number 1 is 0 or doesn't exist
            if pd.isnull(df_extended.iloc[row_num, 1]) or df_extended.iloc[row_num, 1] == 0:
                # Calculate and fill the average value of the column from row=4 to the previous row
                numeric_col = pd.to_numeric(df_extended.iloc[4:last_row_df_extended + 1, col_idx],
                                            errors='coerce').fillna(0)
                avg_value = numeric_col.mean()
                df_extended.loc[last_row_df_extended + 1, col_idx] = avg_value
                break

            # Check if all referenced cells in the formula exist and are not empty
            if any(pd.isnull(row[col_to_idx[re.match(r'([A-Z]+)', cell).group(0)]]) for cell in
                   re.findall(r'[A-Z]+[0-9]+', formula)):
                break
            df_extended.iloc[row_num, col_idx] = apply_formula(row, formula, col_to_idx, df_extended)

    # Create the final DataFrame, including the first four rows of data and processed raw_data
    df_final = pd.concat([df_sample.iloc[:4], df_extended.iloc[4:].reset_index(drop=True)], ignore_index=True)

    # Add new calculated values
    temp_columns = ['Q', 'U', 'V', 'Y', 'BN', 'BO', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CY', 'CZ', 'DA', 'DB', 'DD',
                    'DE', 'DF', 'DG', 'DJ', 'DK', 'DM', 'FB', 'FE']

    pressure_columns = ['AA', 'DC', 'DP', 'DQ', 'FH', 'FI']

    for col in temp_columns:
        if col in col_to_idx:
            idx = col_to_idx[col]
            x = df_final.iloc[last_row_df_extended + 1, idx]
            df_final.iloc[last_row_df_extended + 2, idx] = (x - 32) / 1.8

    for col in pressure_columns:
        if col in col_to_idx:
            idx = col_to_idx[col]
            x = df_final.iloc[last_row_df_extended + 1, idx]
            df_final.iloc[last_row_df_extended + 2, idx] = (x + 14.7) * 6.89476

    # Special calculations
    ju_idx = col_to_idx['JU']
    jv_idx = col_to_idx['JV']
    jw_idx = col_to_idx['JW']
    jt_idx = col_to_idx['JT']
    jz_idx = col_to_idx['JZ']
    jy_idx = col_to_idx['JY']

    df_final.iloc[last_row_df_extended + 2, jv_idx] = df_final.iloc[last_row_df_extended + 1, ju_idx] - df_final.iloc[
        last_row_df_extended + 1, jv_idx]
    df_final.iloc[last_row_df_extended + 2, jw_idx] = df_final.iloc[last_row_df_extended + 1, jw_idx] - df_final.iloc[
        last_row_df_extended + 1, jt_idx]
    df_final.iloc[last_row_df_extended + 2, jz_idx] = df_final.iloc[last_row_df_extended + 1, jz_idx] - df_final.iloc[
        last_row_df_extended + 1, jy_idx]

    # Write specific values
    df_final.iloc[last_row_df_extended + 2, col_to_idx['AF']] = 'Height (ft)'
    df_final.iloc[last_row_df_extended + 4, col_to_idx['BA']] = (df_final.iloc[last_row_df_extended + 1, col_to_idx[
        'BA']] - 32) / 1.8

    # Fill values from AG to BA
    df_final.iloc[last_row_df_extended + 2, col_to_idx['AG']:col_to_idx['BA'] + 1] = [
        5, 16.5, 21.3, 26.8, 32, 43.1, 45.8, 48.3, 53.4, 56, 58.6, 69.9, 72.8, 75, 77.6, 80.2, 82.8, 85.4, 87.4, 96.6,
        104.3
    ]
    df_final.iloc[last_row_df_extended + 3, col_to_idx['AG']:col_to_idx['BA'] + 1] = [
        'TE20101', 'T2', 'T4', 'T6', 'T8', 'T10', 'T11', 'T12', 'T14', 'T15', 'T16', 'T18', 'T19', 'T20', 'T21', 'T22',
        'T23', 'T24', 'T25', 'T26', 'T27'
    ]

    # Fill values from IH to JD
    df_final.iloc[last_row_df_extended + 2, col_to_idx['IH']:col_to_idx['JD'] + 1] = [
        0, 0, 1, 16.5, 21.3, 26.8, 32, 43.1, 45.8, 48.3, 53.4, 56, 58.6, 69.9, 72.8, 75, 77.6, 80.2, 82.8, 85.4, 87.4,
        96.6, 104.3
    ]
    df_final.iloc[last_row_df_extended + 3, col_to_idx['IK']:col_to_idx['JD'] + 1] = [
        'T2', 'T4', 'T6', 'T8', 'T10', 'T11', 'T12', 'T14', 'T15', 'T16', 'T18', 'T19', 'T20', 'T21', 'T22', 'T23',
        'T24', 'T25', 'T26', 'T27'
    ]

    # Fill values from JI to JS
    df_final.iloc[last_row_df_extended + 2, col_to_idx['JI']:col_to_idx['JS'] + 1] = [
        0, 0, 1, 14.9, 20.8, 25.2, 32.1, 38.8, 44, 49.2, 54.3
    ]
    df_final.iloc[last_row_df_extended + 3, col_to_idx['JK']:col_to_idx['JS'] + 1] = [
        'T1', 'T2', 'T4', 'T6', 'T8', 'T10', 'T12', 'T14', 'T16'
    ]

    # Write the processed data to a new Excel file
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df_final.to_excel(writer, sheet_name='Data', index=False, header=False)
        worksheet = writer.sheets['Data']
        worksheet.write('A1', last_row_df_extended)

    # Save df_final for future use
    df_final.to_pickle(df_final_path)

    # Record the value of last_row_df_extended in cell A1 on the calculation worksheet of the address.xlsx file
    address_wb = openpyxl.load_workbook(address_path)
    address_ws = address_wb['calculation']
    address_ws['A1'] = last_row_df_extended
    address_wb.save(address_path)

# Run the functions
save_formulas_to_json('sample.xlsm', 'formulas.json')
process_data('sample.xlsm', 'raw_data.xlsx', 'qiaochu_processed_data.xlsx', 'formulas.json', 'df_final.pkl', 'address.xlsx')
