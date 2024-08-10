import os
import time

from openpyxl import Workbook, load_workbook

import py_aspen
import win32com.client as win32
import numpy as np
import sys
import pandas as pd

import psutil


def get_pid(process_name):
    for proc in psutil.process_iter():
        if proc.name() == process_name:
            return proc.pid


def extract_node_value(value):
    """
    Extracts the node value from the string if it's in the form
    Application.Tree.FindNode("\\Data\\Streams\\GASIN\\Output\\TEMP\\MIXED").
    """
    if isinstance(value, str) and value.startswith("Application.Tree.FindNode("):
        start = value.find("\\")
        end = value.rfind(")")
        return value[start:end].strip('\"')
    return value


def process_value(value, last_row):
    """
    Processes the value to generate the address for the 'qiaochu_processed_data.xlsx' file.
    """
    # Check if value is NaN
    if pd.isna(value):
        return ""

    # Check if value is a float
    if isinstance(value, float):
        print(f"Error: Unexpected float value encountered: {value}")
        return ""

    # Check if the length of value is less than 3
    if len(value) < 3:
        return ""

    # Process specific string format values
    if value[1] == '.' or value[2] == '.':
        parts = value.split('.bot')
        col = parts[0]
        if len(parts) > 1 and '+' in parts[1]:
            offset = int(parts[1].split('+')[1])
            row = last_row + offset + 1
        else:
            row = last_row + 1

        return f"{col}{row}"

    return ""


def get_last_row(filename):
    """
    Finds the last non-zero row in column B of the 'Data' sheet.
    """
    data = pd.read_excel(filename, sheet_name='Data', engine='openpyxl', header=None)
    last_row = data[data.iloc[:, 1] != 0].index[-1] + 1
    return last_row


def get_value_from_address(sheet, address):
    """
    Retrieves the value from the given address in the provided sheet.
    """
    global col_to_idx

    if address:
        # Parse column number
        if len(address) > 1 and address[1].isalpha():
            col_get_val = address[:2]  # The first two characters are letters
            row = int(address[2:]) - 1  # Subtract 1 because iloc is 0-indexed
        else:
            col_get_val = address[0]  # The first character is a letter
            row = int(address[1:]) - 1  # Subtract 1 because iloc is 0-indexed

        # Ensure the column number exists in col_to_idx
        if col_get_val in col_to_idx:
            col_idx = col_to_idx[col_get_val]
            value = sheet.iloc[row, col_idx]
            return value
        else:
            print(f"Error: Column '{col_get_val}' does not exist in the DataFrame.")
            return None

    return None


def get_call_address(filename):
    print(f"Reading initial_cond sheet from {filename}")
    excel_data = pd.read_excel(filename, sheet_name='initial_cond', engine='openpyxl', header=None)
    print("Successfully read initial_cond sheet")

    df_excel_addr_getcall = pd.DataFrame({
        'GASIN_T': [excel_data.iloc[2, 1]],  # B3
        'GASIN_P': [excel_data.iloc[2, 2]],  # C3
        'GASIN_MASS': [excel_data.iloc[2, 3]],  # D3
        'GASIN_H2O': [excel_data.iloc[2, 5]],  # F3
        'GASIN_CO2': [excel_data.iloc[2, 6]],  # G3
        'GASIN_N2': [excel_data.iloc[2, 7]],  # H3
        'GASIN_O2': [excel_data.iloc[2, 8]],  # I3
        'H2OIN_T': [excel_data.iloc[7, 1]],  # B8
        'ABSLEAN_T': [excel_data.iloc[12, 1]],  # B13
        'ABSLEAN_MASS': [excel_data.iloc[12, 3]],  # D13
        'B5_T': [excel_data.iloc[7, 1]],  # B8
        'ABSLEAN_CO2': [excel_data.iloc[12, 4]],  # E13
        'ABSLEAN_H2O': [excel_data.iloc[12, 5]],  # F13
        'ABSLEAN_MEA': [excel_data.iloc[12, 6]],  # G13
        'ABSLEAN_LOADING': [excel_data.iloc[12, 7]],  # H13
        'ABSORBER_PD1': [excel_data.iloc[18, 1]],  # B19
        'ABSORBER_PD2': [excel_data.iloc[18, 2]],  # C19
        'ABSORBER_PD3': [excel_data.iloc[18, 3]],  # D19
        'ABSORBER_PD4': [excel_data.iloc[18, 4]],  # E19
        'ABSORBER_PD5': [excel_data.iloc[18, 5]],  # F19
        'ABSORBER_PST1': [excel_data.iloc[18, 6]],  # G19
        'FLASH_T': [excel_data.iloc[36, 1]],  # B37
        'LEANCOOL_T': [excel_data.iloc[45, 2]],  # C46
        'PUMP2_P': [excel_data.iloc[56, 1]],  # B57
        'WASH_PST1': [excel_data.iloc[64, 1]]  # B65
    })

    print("Created df_excel_addr_getcall")

    last_row = get_last_row('qiaochu_processed_data.xlsx')
    print(f"Last row in qiaochu_processed_data.xlsx: {last_row}")

    df_excel_input_getcall = pd.DataFrame(columns=df_excel_addr_getcall.columns)
    for key, value in df_excel_addr_getcall.iloc[0].items():
        processed_value = process_value(value, last_row)
        df_excel_input_getcall.at[0, key] = processed_value

    print("Created df_excel_input_getcall")

    df_excel_in_value_getcall = pd.DataFrame(columns=df_excel_addr_getcall.columns)
    processed_data = pd.read_excel('qiaochu_processed_data.xlsx', sheet_name='Data', engine='openpyxl', header=None)
    for key, address in df_excel_input_getcall.iloc[0].items():
        actual_value = get_value_from_address(processed_data, address)
        df_excel_in_value_getcall.at[0, key] = actual_value

    print("Created df_excel_in_value_getcall")

    df_excel_in_value_getcall['GASIN_MASS'] = df_excel_in_value_getcall['GASIN_MASS'] * 0.453592
    print("Processed GASIN_MASS")

    T = df_excel_in_value_getcall['GASIN_T'].iloc[0]
    P = df_excel_in_value_getcall['GASIN_P'].iloc[0]
    df_excel_in_value_getcall['GASIN_H2O'] = df_excel_in_value_getcall['GASIN_H2O'] / 100 * 0.61078 * np.exp(
        (17.27 * T) / (T + 237.3)) / P
    print("Processed GASIN_H2O")

    h2o = df_excel_in_value_getcall['GASIN_H2O']
    df_excel_in_value_getcall['GASIN_CO2'] = (1 - h2o) * df_excel_in_value_getcall['GASIN_CO2'] / 100
    df_excel_in_value_getcall['GASIN_O2'] = (1 - h2o) * df_excel_in_value_getcall['GASIN_O2'] / 100
    co2 = df_excel_in_value_getcall['GASIN_CO2']
    o2 = df_excel_in_value_getcall['GASIN_O2']
    df_excel_in_value_getcall['GASIN_N2'] = 1 - h2o - co2 - o2
    print("Processed GASIN components")

    df_excel_in_value_getcall['ABSLEAN_MASS'] = df_excel_in_value_getcall['ABSLEAN_MASS'] * 0.453592
    print("Processed ABSLEAN_MASS")

    loading = df_excel_in_value_getcall['ABSLEAN_LOADING']
    x = df_excel_in_value_getcall['ABSLEAN_MEA']
    df_excel_in_value_getcall['ABSLEAN_MEA'] = (1 + loading + (61.08 / 18.02) * (100 / x - 1)) ** (-1)
    print("Processed ABSLEAN_MEA")

    MEA = df_excel_in_value_getcall['ABSLEAN_MEA']
    df_excel_in_value_getcall['ABSLEAN_CO2'] = loading * MEA
    CO2 = df_excel_in_value_getcall['ABSLEAN_CO2']
    df_excel_in_value_getcall['ABSLEAN_H2O'] = 1 - CO2 - MEA
    print("Processed ABSLEAN components")

    df_excel_in_value_getcall['ABSORBER_PST1'] = df_excel_in_value_getcall['GASIN_P'] - (
            df_excel_in_value_getcall['ABSORBER_PD1'] + df_excel_in_value_getcall['ABSORBER_PD2'] +
            df_excel_in_value_getcall['ABSORBER_PD3'] + df_excel_in_value_getcall['ABSORBER_PD4'] +
            df_excel_in_value_getcall['ABSORBER_PD5']) * 0.249
    print("Processed ABSORBER_PST1")

    df_excel_in_value_getcall['B5_T'] = df_excel_in_value_getcall['H2OIN_T']
    df_excel_in_value_getcall['LEANCOOL_T'] = df_excel_in_value_getcall['ABSLEAN_T']
    print("Processed temperatures")

    df_excel_in_value_getcall['PUMP2_P'] = df_excel_in_value_getcall['PUMP2_P'] * 6.89476 + 101.315
    print("Processed PUMP2_P")

    df_aspen_in = pd.DataFrame({
        'GASIN_T': [extract_node_value(excel_data.iloc[4, 1])],  # B5
        'GASIN_P': [extract_node_value(excel_data.iloc[4, 2])],  # C5
        'GASIN_MASS': [extract_node_value(excel_data.iloc[4, 3])],  # D5
        'GASIN_H2O': [extract_node_value(excel_data.iloc[4, 5])],  # F5
        'GASIN_CO2': [extract_node_value(excel_data.iloc[4, 6])],  # G5
        'GASIN_N2': [extract_node_value(excel_data.iloc[4, 7])],  # H5
        'GASIN_O2': [extract_node_value(excel_data.iloc[4, 8])],  # I5
        'H2OIN_T': [extract_node_value(excel_data.iloc[9, 1])],  # B10
        'ABSLEAN_T': [extract_node_value(excel_data.iloc[14, 1])],  # B15
        'ABSLEAN_CO2': [extract_node_value(excel_data.iloc[14, 4])],  # E15
        'ABSLEAN_H2O': [extract_node_value(excel_data.iloc[14, 5])],  # F15
        'ABSLEAN_MEA': [extract_node_value(excel_data.iloc[14, 6])],  # G15
        'ABSORBER_PST1': [extract_node_value(excel_data.iloc[20, 6])],  # G21
        'B5_T': [extract_node_value(excel_data.iloc[27, 1])],  # B28
        'FLASH_T': [extract_node_value(excel_data.iloc[38, 1])],  # B39
        'LEANCOOL_T': [extract_node_value(excel_data.iloc[46, 1])],  # B47
        'PUMP2_P': [extract_node_value(excel_data.iloc[58, 1])],  # B59
        'WASH_PST1': [extract_node_value(excel_data.iloc[65, 1])],  # B66
        'ABSLEAN_MASS': [extract_node_value(excel_data.iloc[14, 3])]  # D15
    })

    print("Created df_aspen_in")

    export_data_getcall = pd.read_excel(filename, sheet_name='data_export', engine='openpyxl', header=None)
    print("Read data_export sheet")

    df_aspen_out_getcall = pd.DataFrame({
        'REGEN_LIQ_TOP': [extract_node_value(export_data_getcall.iloc[9, 13])],  # N10
        'REGEN_LIQ_BOT': [extract_node_value(export_data_getcall.iloc[11, 13])],  # N12
        'ABSORBER_LIQ_TOP': [extract_node_value(export_data_getcall.iloc[19, 13])],  # N20
        'ABSORBER_LIQ_BOT': [extract_node_value(export_data_getcall.iloc[21, 13])],  # N22
    })

    ex_keys = export_data_getcall.iloc[32:68, 3].tolist()  # D33-D68
    ex_values = [extract_node_value(val) for val in export_data_getcall.iloc[32:68, 4].tolist()]  # E33-E68

    # Convert the read keys and values to a dictionary and create a new DataFrame
    data_dict = {key: [value] for key, value in zip(ex_keys, ex_values)}
    ex_df = pd.DataFrame(data_dict)

    # Add the new DataFrame to the existing df_aspen_out_getcall
    df_aspen_out_getcall = pd.concat([df_aspen_out_getcall, ex_df], axis=1)

    print("Created df_aspen_out_getcall")

    # Create a dictionary to store all new column data
    new_columns = {}
    # Store REGEN_LIQ column data
    for j in range(1, 42):
        column_name = f'REGEN_LIQ_{j}'
        column_value = extract_node_value((export_data_getcall.iloc[10, 13]).rsplit('\\', 1)[0] + f'\\{j}' + "\")")
        new_columns[column_name] = column_value

    # Store ABSORBER_LIQ column data
    for j in range(1, 91):
        column_name = f'ABSORBER_LIQ_{j}'
        column_value = extract_node_value((export_data_getcall.iloc[20, 13]).rsplit('\\', 1)[0] + f'\\{j}' + "\")")
        new_columns[column_name] = column_value

    # Convert the dictionary to a DataFrame
    new_columns_df = pd.DataFrame([new_columns])

    # Merge the new DataFrame with the original df_aspen_out_getcall
    df_aspen_out_getcall = pd.concat([df_aspen_out_getcall, new_columns_df], axis=1)

    print("Processed additional REGEN_LIQ and ABSORBER_LIQ columns")

    return df_excel_addr_getcall, df_aspen_in, df_aspen_out_getcall, df_excel_input_getcall, df_excel_in_value_getcall


def import_initial_conditions(df_excel_in_value_im: pd.DataFrame, df_aspen_in_im: pd.DataFrame):
    global pyaspen

    columns_excel = set(df_excel_in_value_im.columns)
    columns_aspen = set(df_aspen_in_im.columns)
    common_column = columns_excel.intersection(columns_aspen)
    missing_columns = columns_excel - columns_aspen
    exclude_columns = {'ABSORBER_PD2', 'ABSORBER_PD1', 'ABSLEAN_LOADING', 'ABSORBER_PD4', 'ABSORBER_PD5',
                       'ABSORBER_PD3'}
    filtered_missing_columns = missing_columns - exclude_columns
    if filtered_missing_columns:
        raise KeyError(f"The following columns are missing in df_aspen_in_im: {filtered_missing_columns}")
    nodes = []
    values = []
    call_address = {}
    for column in common_column:
        nodes.append(column)
        values.append(df_excel_in_value_im.at[0, column])
        address = df_aspen_in_im.at[0, column]
        call_address[column] = address

    pyaspen.assign_node_values(nodes, values, call_address)


def check_convergence(stream_in_add_che, stream_back_add_che):
    global pyaspen
    global aspen_runs
    global last_error
    global divergence

    last_error = "error"
    i0 = 0  # run times
    while last_error == "error":
        if aspen_runs == 1:
            flowrate0 = pyaspen.get_target_value1(ABSLEAN_FR) - diver0
            pyaspen.assign_node_value1(flowrate0, ABSLEAN_FR)

        pyaspen.run_simulation(reinit=True)
        """
        if i0 == 0:
            pyaspen.run_simulation(reinit=False)
        else:
            pyaspen.run_simulation(reinit=True)
        i0 = 1
        """
        print("Ran Aspen simulation in convergence check")
        aspen_runs += 1
        print(f"Total counts: {aspen_runs}")
        last_error = pyaspen.result_error()
        print(last_error)
        if last_error == "error":
            flowrate = pyaspen.get_target_value1(ABSLEAN_FR) - 100
            pyaspen.assign_node_value1(flowrate, ABSLEAN_FR)
            divergence += 100
            print(f"ABSLEAN flowrate: {flowrate}, divergence: {divergence}, divergence0: {diver0}")

    stream_in_values = pyaspen.get_target_value1(stream_in_add_che)
    stream_back_values = pyaspen.get_target_value1(stream_back_add_che)
    print(np.array(stream_back_values))
    print(np.array(stream_in_values))
    relative_error = np.abs((np.array(stream_back_values) - np.array(stream_in_values)) / np.array(stream_in_values))
    rmsre = np.sqrt(np.mean(relative_error ** 2))
    print("relative error RMSRE of H2O in Stream S6")
    print(rmsre)
    return rmsre


def run_aspen(t_h2oin: int, aspen_file: str, direct: str, target_temp: float,
              duty_range: (float, float), num_points: int):
    global divergence
    global pyaspen
    global aspen_runs
    global last_error

    duties = np.linspace(duty_range[0], duty_range[1], num_points)
    closest_duty = None
    smallest_diff = float('inf')

    for duty in duties:
        pyaspen.assign_node_value1(duty, r"\Data\Blocks\ABSORBER\Input\HEATER_DUTY\25")
        last_error = "error"
        i1 = 0
        while last_error == "error":
            pyaspen.run_simulation(reinit=True)
            """
            if i1 == 0:
                pyaspen.run_simulation(reinit=False)
            else:
                pyaspen.run_simulation(reinit=True)
            i1 = 1
            """
            print("Ran Aspen simulation in duty loop")
            aspen_runs += 1
            print(f"Total counts: {aspen_runs}")
            last_error = pyaspen.result_error()
            print(last_error)
            if last_error == "error":
                flowrate = pyaspen.get_target_value1(ABSLEAN_FR) - 100
                pyaspen.assign_node_value1(flowrate, ABSLEAN_FR)
                divergence += 100
                print(f"ABSLEAN flowrate: {flowrate}, divergence: {divergence}, divergence0: {diver0}")

        print(f"intercooler duty:")
        print(duty)
        last_error = pyaspen.result_error()
        print(last_error)
        current_temp = pyaspen.get_target_value1(r"\Data\Blocks\ABSORBER\Output\TLIQ\25")
        print("input temp for H2OIN:")
        print(t_h2oin)
        print("absorber stage 25 temp:")
        print(current_temp)
        print("absorber stage 26 temp:")
        print(pyaspen.get_target_value1(r"\Data\Blocks\ABSORBER\Output\TLIQ\26"))
        diff = abs(current_temp - target_temp)
        if diff < smallest_diff and last_error == 'OK':
            smallest_diff = diff
            closest_duty = duty
            print(f"smallest difference of temperature for Stage 25: {smallest_diff}")

        # restart
        if round(aspen_runs) % 30 == 0:
            print('RESTART ASPEN')
            for proc_run in psutil.process_iter():
                if any(procstr in proc_run.name() for procstr in ['AspenPlus', 'Aspen', 'apmain']):
                    print(f'Killing {proc_run.name()}')
                    proc_run.kill()
                    time.sleep(2)

            pyaspen = py_aspen.PyASPENPlus()
            pyaspen.init_app("11.0")
            print("Initialized Aspen application")
            pyaspen.load_ap_file(aspen_file, direct)
            import_initial_conditions(df_excel_in_value, df_aspen_in)
            flowrate = pyaspen.get_target_value1(ABSLEAN_FR) - divergence - diver0
            pyaspen.assign_node_value1(flowrate, ABSLEAN_FR)
            pyaspen.assign_node_value1(t_h2oin, r'\Data\Streams\H2OIN\Input\TEMP\MIXED')

    pyaspen.assign_node_value1(closest_duty, r"\Data\Blocks\ABSORBER\Input\HEATER_DUTY\25")
    pyaspen.run_simulation()
    """
    if not last_error:
        pyaspen.run_simulation()
    else:
        pyaspen.run_simulation(reinit=False)
    """
    last_error = pyaspen.result_error()
    return smallest_diff


def export_data(df_aspen_out_exp: pd.DataFrame):
    global pyaspen

    df_aspen_out_value = {}
    for key in df_aspen_out_exp.columns:
        address = df_aspen_out_exp[key].iloc[0]
        df_aspen_out_value[key] = pyaspen.get_target_value1(str(address))

    df_aspen_out_value = pd.DataFrame.from_dict(df_aspen_out_value, orient='index', columns=['Value'])

    file_name_exp = 'qiaochu_processed_data.xlsx'
    try:
        book = load_workbook(file_name_exp)
    except FileNotFoundError:
        book = Workbook()

    if 'Model' in book.sheetnames:
        sheet = book['Model']
    else:
        sheet = book.create_sheet('Model')

    sheet['D4'] = 'total packing'
    sheet['E4'] = 85.4  # total height
    sheet['F4'] = 'ft'
    sheet['E5'] = 0.948889  # distance between 2 stages
    sheet['F8'] = 'stage'
    sheet['G8'] = 'liquid temperature'
    sheet['L9'] = 'Height(ft)'
    sheet['M9'] = 'T/C'
    sheet['G9'] = 'C'
    sheet['H9'] = 'C'
    sheet['E10'] = 101.9

    for i in range(90):
        sheet[f'F{i + 10}'] = i + 1

    sheet['L10'] = 104.3

    for i in range(90):
        height_value = 101.9 - 0.948889 * (i + 1)
        sheet[f'E{i + 11}'] = height_value
        sheet[f'G{i + 10}'] = df_aspen_out_value.loc[f'ABSORBER_LIQ_{i + 1}', 'Value']
        sheet[f'L{i + 11}'] = height_value
        sheet[f'M{i + 11}'] = df_aspen_out_value.loc[f'ABSORBER_LIQ_{i + 1}', 'Value']

    sheet['E100'] = 0
    sheet['L101'] = 0
    sheet['M10'] = df_aspen_out_value.loc['ABSORBER_LIQ_TOP', 'Value']
    sheet['M101'] = df_aspen_out_value.loc['ABSORBER_LIQ_BOT', 'Value']

    sheet['D115'] = 'total packing'
    sheet['E115'] = 39.6837  # total height
    sheet['F115'] = 'ft'
    sheet['E116'] = 0.967895122  # distance between 2 stages
    sheet['F118'] = 'stage'
    sheet['G118'] = 'liquid temperature'
    sheet['L119'] = 'Height(ft)'
    sheet['M119'] = 'T/C'
    sheet['G119'] = 'C'
    sheet['H119'] = 'C'
    sheet['E120'] = 54.3

    for i in range(41):
        sheet[f'F{i + 120}'] = i + 1

    sheet['L120'] = 55

    for i in range(41):
        height_value = 54.3 - 0.967895122 * (i + 1)
        sheet[f'E{i + 121}'] = height_value
        sheet[f'G{i + 120}'] = df_aspen_out_value.loc[f'REGEN_LIQ_{i + 1}', 'Value']
        sheet[f'L{i + 121}'] = height_value
        sheet[f'M{i + 121}'] = df_aspen_out_value.loc[f'REGEN_LIQ_{i + 1}', 'Value']

    sheet['E161'] = 0
    sheet['L162'] = 0
    sheet['M120'] = df_aspen_out_value.loc['REGEN_LIQ_TOP', 'Value']
    sheet['M162'] = df_aspen_out_value.loc['REGEN_LIQ_BOT', 'Value']

    start_row = 180
    col_name = 'D'
    col_value = 'E'

    for i, index in enumerate(df_aspen_out_value.index):
        if 'LIQ' not in index:
            sheet[f'{col_name}{start_row + i}'] = index
            sheet[f'{col_value}{start_row + i}'] = df_aspen_out_value.loc[index, 'Value']

    book.save(file_name_exp)


col_to_idx = {}
for i in range(1, 703):  # 1 to 702 (corresponding to 'A' to 'ZZ')
    col = ''
    n = i
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        col = chr(65 + remainder) + col
    col_to_idx[col] = i - 1

print("Column index mapping created successfully.")
ABSLEAN_FR = r"\Data\Streams\ABSLEAN\Input\TOTFLOW\MIXED"

file_name = r'address.xlsx'
aspen_runs = 1
last_error = 'OK'
df_excel_addr, df_aspen_in, df_aspen_out, df_excel_input, df_excel_in_value = get_call_address(file_name)
print("Completed get_call_address function")

aspen_file_name = 'co2_cap_mea.bkp'
file_dir = os.getcwd()

pyaspen = py_aspen.PyASPENPlus()
pyaspen.init_app("11.0")
print("Initialized Aspen application")

pyaspen.load_ap_file(aspen_file_name, file_dir)
print("Loaded Aspen file")
pyaspen.run_simulation()
import_initial_conditions(df_excel_in_value, df_aspen_in)

count = 0  # loop number
divergence = 0
diver0 = 0  # initial ABSLEAN flowrate reduction guess for simulation convergence

stream_in_add = r'\Data\Streams\ABSLEAN\Output\MOLEFLOW\MIXED\H2O'
stream_out_add = r'\Data\Streams\S6\Output\MOLEFLOW\MIXED\H2O'
error1 = 0.01  # relative error for H2O convergence allowed
gasin_temp = pyaspen.get_target_value1(r'\Data\Streams\GASIN\Input\TEMP\MIXED')
while check_convergence(stream_in_add, stream_out_add) > error1:
    count += 1
    print(f"Ran Aspen simulation with loop {count} start!")
    if count > 20:
        print(f"The simulation does not converge as expected, but has run {count - 1} loops")
        break
    T_H2Oin = (gasin_temp - 10 + count * 1)
    pyaspen.assign_node_value1(T_H2Oin, r'\Data\Streams\H2OIN\Input\TEMP\MIXED')
    print(f"Start simulate H2OIN temp = {T_H2Oin}")
    print("assign node value finished")
    small_diff = run_aspen(T_H2Oin, aspen_file_name, file_dir, 40, (-35, -5), 5)
    print(
        f"Ran Aspen simulation with loop {count} finished, H2OIn has a temperature of {T_H2Oin}, the intercooling temperature difference is {small_diff}")

export_data(df_aspen_out)
print("Exported data from Aspen to Excel")

pyaspen.close_app()
print("Closed Aspen application")

for proc in psutil.process_iter():
    if any(procstr in proc.name() for procstr in ['AspenPlus', 'Aspen', 'apmain']):
        print(f'Killing {proc.name()}')
        proc.kill()
        time.sleep(2)
print("Terminated AspenPlus process")
